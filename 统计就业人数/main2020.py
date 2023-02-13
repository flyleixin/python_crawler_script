import openpyxl

bsgs = ["北京市", "深圳市", "广州市", "上海市"]
jz = ["浙江省", "江苏省"]
jx = ["江西省"]


def excelData():
    wb = openpyxl.load_workbook('./2020.xlsx')
    sheet = wb.active
    minrow = sheet.min_row  # 最小行
    maxrow = sheet.max_row  # 最大行
    mincol = sheet.min_column  # 最小列
    maxcol = sheet.max_column  # 最大列
    """
    # 按行读取
    for i in range(minrow, maxrow + 1):
        for j in range(mincol, maxcol + 1):
            cell = sheet.cell(i, j).value
            print(cell, end=" ")
        print()

    # cells = sheet['B']
    # for i in cells:
    #     print(i.value)
    """
    # 按行读取
    for i in range(minrow, maxrow + 1):
        for j in range(mincol, maxcol + 1):
            cell = sheet.cell(i, j).value
            # print(cell, end=" ")
        # print()
    hang = maxrow - (minrow + 1) + 1
    print("总人数为: %d" % hang)

    cells = sheet['AN']
    shiye = 0
    for i in cells:
        if i.value == "701":
            shiye += 1
    jiuye = (hang - shiye) / hang
    print("就业率: %f" % jiuye)
    yanjiusheng = 0
    for i in cells:
        if i.value == "800" or i.value == "850":
            yanjiusheng += 1
    print("研究生录取数 %d" % yanjiusheng)

    gouqi = 0
    cells = sheet['AS']
    for i in cells:
        if i.value == "31" or i.value == "20" or i.value == "29":
            gouqi += 1
    print("国企 %d" % gouqi)

    siqi = 0
    for i in cells:
        if i.value == "39" or i.value == "32":
            siqi += 1
    print("私企 %d" % siqi)

    cells = sheet["AX"]
    bsgs_tot = 0
    jz_tot = 0
    jx_tot = 0
    for i in cells:
        for j in bsgs:
            if j in i.value:
                bsgs_tot += 1
        for j in jz:
            if j in i.value:
                jz_tot += 1
        for j in jx:
            if j in i.value:
                jx_tot += 1

    print("北上广深: %d" % bsgs_tot)
    print("江浙 %d" % jz_tot)
    print("江西本地 %d" % jx_tot)
    print("------------")


def excelData_zy():
    wb = openpyxl.load_workbook('./2020.xlsx')
    sheet = wb.active
    minrow = sheet.min_row  # 最小行
    maxrow = sheet.max_row  # 最大行
    mincol = sheet.min_column  # 最小列
    maxcol = sheet.max_column  # 最大列
    # print(minrow)
    # print(maxrow)
    zy_list = sheet['Q']
    se = set()
    for i in zy_list:
        se.add(i.value)

    # 遍历set
    for s in se:
        print(s)

        zy_peo_tot = 0
        zy_shiye = 0
        zy_yanjiusheng = 0
        zy_gouqi = 0
        zy_siqi = 0
        zy_bsgs_tot = 0
        zy_jz_tot = 0
        zy_jx_tot = 0
        zy_budui = 0
        zy_yiliao = 0
        zy_zhongchujiaoyu = 0
        # 按行读取
        for i in range(minrow, maxrow + 1):
            col_zy = sheet.cell(i, 17).value
            # 专业对上
            if col_zy == s:
                zy_peo_tot += 1
                if sheet.cell(i, 40).value == "701":
                    zy_shiye += 1
                if sheet.cell(i, 40).value == "800" or sheet.cell(i, 40).value == "850":
                    zy_yanjiusheng += 1
                if sheet.cell(i, 45).value == "31" or sheet.cell(i, 45).value == "20" or sheet.cell(i,
                                                                                                    45).value == "29":
                    zy_gouqi += 1
                if sheet.cell(i, 45).value == "39" or sheet.cell(i, 45).value == "32":
                    zy_siqi += 1
                if sheet.cell(i, 45).value == '40':
                    zy_budui += 1
                if sheet.cell(i, 45).value == '23':
                    zy_yiliao += 1
                if sheet.cell(i, 45).value == '22':
                    zy_zhongchujiaoyu += 1
                diqu = sheet.cell(i, 50).value
                for k in bsgs:
                    if k in diqu:
                        zy_bsgs_tot += 1
                for k in jz:
                    if k in diqu:
                        zy_jz_tot += 1
                for k in jx:
                    if k in diqu:
                        zy_jx_tot += 1
        print("总人数为: %d" % zy_peo_tot)
        jiuye = (zy_peo_tot - zy_shiye) / zy_peo_tot
        print("就业率: %f" % jiuye)
        print("研究生录取数 %d" % zy_yanjiusheng)
        print("国企 %d" % zy_gouqi)
        print("私企 %d" % zy_siqi)
        print("北上广深: %d" % zy_bsgs_tot)
        print("江浙 %d" % zy_jz_tot)
        print("江西本地 %d" % zy_jx_tot)
        print("部队 %d" % zy_budui)
        print("医疗卫生单位 %d" % zy_yiliao)
        print("中初教育单位 %d" % zy_zhongchujiaoyu)

        print("------------")


if __name__ == '__main__':
    excelData()
    excelData_zy()
